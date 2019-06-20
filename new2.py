import urllib.request
from urllib.request import urlretrieve
import urllib.parse
import time
from tqdm import tqdm
import os
import ssl
import re
from bs4 import BeautifulSoup

from openpyxl.workbook import Workbook
import openpyxl
from openpyxl import load_workbook

# Bypass SSL verification
ssl._create_default_https_context = ssl._create_unverified_context
CONTEXT = ssl._create_unverified_context()

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.90 Safari/537.36'
}

links = open('sitemap.txt', 'r')
link_id = 0

for link in links:
    link_id = link_id+1
    print("**************************")
    print('LINK ID:', link_id)

    link = link.strip()
    headers = {}
    req = urllib.request.Request(link, headers=HEADERS)
    try:
        resp = urllib.request.urlopen(req, timeout=60, context=CONTEXT)

        respData = resp.read()
        resp.close()
        print('Done')

        soup = BeautifulSoup(respData, "html.parser")

        sku_src = soup.find('div', {'class': 'prod-buy__article'})
        if sku_src:
            sku = sku_src.find('span').text
            print('SKU:', sku)

            wb = openpyxl.load_workbook(filename='products.xlsx')
            ws = wb.active

            loop = True

            # start and end numbers of rows in the xlsx file
            # the last one not included
            for row in range(2, 9):
                if loop is True:
                    # for choose column D:
                    for column in range(4, 5):
                        xl_cell = ws.cell(row=row, column=column).value
                        print("xl_cell: ", xl_cell, "row: ", row)
                        
                        
                        if re.sub("\D", "", str(sku)) != re.sub("\D", "", str(xl_cell)):
                            loop = True
                            print("not match (((")
                        else:    
                            print("equals!!!")
                            loop = False
                            

                            dsc_src = soup.find('div', {'class': 'tabs__chars'})
                            if dsc_src:
        
                                dsc = str(dsc_src)
                                print('DSC: ', dsc)
                                ws.cell(row=row, column=29).value = dsc

                                wb.save('products.xlsx')

        continue
    except urllib.error.URLError as e:
        print(e.reason)
